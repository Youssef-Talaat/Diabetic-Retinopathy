import numpy as np
from keras.layers import Dense, AveragePooling2D, GlobalAveragePooling2D, Dropout, Flatten
from keras.optimizers import SGD
from keras.applications import VGG16, VGG19, ResNet50, InceptionV3
from keras.models import Model
from sklearn.metrics import confusion_matrix
import keras
import cv2
import tensorflow as tf
import openpyxl
import datetime


def getSheets(path):
    print("getSheets(path)")
    workbook = openpyxl.load_workbook(path)
    worksheet = []
    for i in range(classes_num):
        worksheet.append(workbook["Class" + str(i)])
    return worksheet


def getNumOfImagesPerClass():
    print("getNumOfImagesPerClass()")
    for i in range(classes_num):
        images_num_per_class_original[i] = min(target_images_per_class, ws_original[i].max_row)
        images_num_per_class_sampling[i] = target_images_per_class - images_num_per_class_original[i]


def readImagesToArrays():
    print("readImagesToArrays()")
    temp0 = np.zeros(((images_num_per_class_original[0] + images_num_per_class_sampling[0]), image_dimension, image_dimension, 3), dtype='uint8')
    temp1 = np.zeros(((images_num_per_class_original[1] + images_num_per_class_sampling[1]), image_dimension, image_dimension, 3), dtype='uint8')
    temp2 = np.zeros(((images_num_per_class_original[2] + images_num_per_class_sampling[2]), image_dimension, image_dimension, 3), dtype='uint8')
    temp3 = np.zeros(((images_num_per_class_original[3] + images_num_per_class_sampling[3]), image_dimension, image_dimension, 3), dtype='uint8')
    temp4 = np.zeros(((images_num_per_class_original[4] + images_num_per_class_sampling[4]), image_dimension, image_dimension, 3), dtype='uint8')
    images = np.array([temp0, temp1, temp2, temp3, temp4])

    for i in range(classes_num):
        counter = 0
        print("1st Image Name Original", ws_original[i].cell(1, 1).value)
        print("1st Image Name Sampling", ws_sampling[i].cell(1, 1).value)
        for j in range(images_num_per_class_original[i]):
            images[i][counter] = cv2.imread(original_images_path + str(ws_original[i].cell(j+1, 1).value))
            counter += 1
        for j in range(images_num_per_class_sampling[i]):
            images[i][counter] = cv2.imread(sampling_images_path + str(ws_sampling[i].cell(j+1, 1).value))
            counter += 1

    print(images_num_per_class_original)
    print(images_num_per_class_sampling)
    return images


def splitData():
    print("splitData()")
    for c in range(classes_num):
        train_size_per_class_original[c] = int(images_num_per_class_original[c] * train_percentage + 0.5)
        test_size_per_class_original[c] = images_num_per_class_original[c] - train_size_per_class_original[c]
        print(str(train_size_per_class_original[c]) + " , " + str(test_size_per_class_original[c]))

        train_size_per_class_sampling[c] = int(images_num_per_class_sampling[c] * train_percentage + 0.5)
        test_size_per_class_sampling[c] = images_num_per_class_sampling[c] - train_size_per_class_sampling[c]
        print(str(train_size_per_class_sampling[c]) + " , " + str(test_size_per_class_sampling[c]))

    classes_train_images_local_original = np.array([classes_images[0][0:train_size_per_class_original[0]],
                                                    classes_images[1][0:train_size_per_class_original[1]],
                                                    classes_images[2][0:train_size_per_class_original[2]],
                                                    classes_images[3][0:train_size_per_class_original[3]],
                                                    classes_images[4][0:train_size_per_class_original[4]]])
    classes_test_images_local_original = np.array([classes_images[0][train_size_per_class_original[0]:images_num_per_class_original[0]],
                                                   classes_images[1][train_size_per_class_original[1]:images_num_per_class_original[1]],
                                                   classes_images[2][train_size_per_class_original[2]:images_num_per_class_original[2]],
                                                   classes_images[3][train_size_per_class_original[3]:images_num_per_class_original[3]],
                                                   classes_images[4][train_size_per_class_original[4]:images_num_per_class_original[4]]])

    classes_train_images_local_sampling = np.array([classes_images[0][images_num_per_class_original[0]:(images_num_per_class_original[0] + train_size_per_class_sampling[0])],
                                                    classes_images[1][images_num_per_class_original[1]:(images_num_per_class_original[1] + train_size_per_class_sampling[1])],
                                                    classes_images[2][images_num_per_class_original[2]:(images_num_per_class_original[2] + train_size_per_class_sampling[2])],
                                                    classes_images[3][images_num_per_class_original[3]:(images_num_per_class_original[3] + train_size_per_class_sampling[3])],
                                                    classes_images[4][images_num_per_class_original[4]:(images_num_per_class_original[4] + train_size_per_class_sampling[4])]])
    classes_test_images_local_sampling = np.array([classes_images[0][(images_num_per_class_original[0] + train_size_per_class_sampling[0]):],
                                                   classes_images[1][(images_num_per_class_original[1] + train_size_per_class_sampling[1]):],
                                                   classes_images[2][(images_num_per_class_original[2] + train_size_per_class_sampling[2]):],
                                                   classes_images[3][(images_num_per_class_original[3] + train_size_per_class_sampling[3]):],
                                                   classes_images[4][(images_num_per_class_original[4] + train_size_per_class_sampling[4]):]])

    return classes_train_images_local_original, classes_test_images_local_original, classes_train_images_local_sampling, classes_test_images_local_sampling


def writeToMainArrays(distribution_type):
    train_counter = 0
    test_counter = 0
    if distribution_type == "block":
        '''for i in range(classes_num):
            for j in range(train_size_per_class[i]):
                train_images[train_counter] = classes_train_images[i][j]
                train_labels[train_counter] = i
                train_counter += 1
            for k in range(test_size_per_class[i]):
                test_images[test_counter] = classes_test_images[i][k]
                test_labels[test_counter] = i
                test_counter += 1'''
    elif distribution_type == "batch":
        pass
    elif distribution_type == "01234":
        for i in range(classes_num):
            train_counter = i
            test_counter = i
            for j in range(train_size_per_class_original[i]):
                train_images[train_counter] = classes_train_images_original[i][j]
                train_labels[train_counter] = i
                train_counter += classes_num
            for j in range(train_size_per_class_sampling[i]):
                train_images[train_counter] = classes_train_images_sampling[i][j]
                train_labels[train_counter] = i
                train_counter += classes_num
            for j in range(test_size_per_class_original[i]):
                test_images[test_counter] = classes_test_images_original[i][j]
                test_labels[test_counter] = i
                test_counter += classes_num
            for j in range(test_size_per_class_sampling[i]):
                test_images[test_counter] = classes_test_images_sampling[i][j]
                test_labels[test_counter] = i
                test_counter += classes_num


def checkArrays():
    print("checkArrays()")
    zeros_array = np.zeros((image_dimension, image_dimension, 3), dtype='uint8')
    total_train = sum(train_size_per_class_original) + sum(train_size_per_class_sampling)
    total_test = sum(test_size_per_class_original) + sum(test_size_per_class_sampling)
    zero_train_counter = 0
    zero_test_counter = 0
    for i in range(total_train):
        if np.array_equal(train_images[i], zeros_array):
            zero_train_counter += 1
    for i in range(total_test):
        if np.array_equal(test_images[i], zeros_array):
            zero_test_counter += 1

    print("Zero Images in Train = ", zero_train_counter)
    print("Zero Images in Test = ", zero_test_counter)


'''def createModel():
    print("createModel()")
    baseModel = VGG16(include_top=False, weights=weights, input_tensor=None, input_shape=(image_dimension, image_dimension, 3))

    x = baseModel.output
    x = GlobalAveragePooling2D()(x)
    x = Dropout(0.5)(x)
    x = Dense(256, activation='relu', kernel_regularizer='l2')(x)
    x = Dropout(0.5)(x)
    headModel = Dense(classes_num, activation='softmax', kernel_regularizer='l2')(x)

    model = Model(inputs=baseModel.input, outputs=headModel)

    for layer in baseModel.layers:
        layer.trainable = False

    return model


def trainModel(_model, _epochs, _train_2_times):
    print("trainModel(model, epochs, _train_2_times)")
    _model.compile(optimizer=optimizer, loss=loss_function, metrics=[metrics])
    _model.summary()
    _model.fit(train_images, train_labels, batch_size=batch_size, epochs=_epochs)

    if _train_2_times:
        for layer in model.layers[:15]:
            layer.trainable = False
        for layer in model.layers[15:]:
            layer.trainable = True

        _model.compile(optimizer=SGD(lr=0.0001, momentum=0.9), loss=loss_function, metrics=[metrics])
        _model.summary()
        _model.fit(train_images, train_labels, batch_size=batch_size, epochs=_epochs)

    return model


def evaluateModel(_model):
    print("evaluateModel(model)")
    model.save('Stages_5Classes_fullModel.h5')
    model.save_weights('Stages_5Classes_weights.h5')

    test_loss, test_acc = _model.evaluate(test_images, test_labels)

    print('\nTest accuracy:', test_acc)
    print('\nTest loss:', test_loss)

    pred = _model.predict(test_images)
    print(pred)
    max_predictions = tf.argmax(pred, axis=1)
    print(max_predictions)
    print(tf.math.confusion_matrix(labels=test_labels, predictions=max_predictions))

    print('Sum Test: ', sum(sum(pred)))'''


def calculateTime():
    print("calculateTime()")
    end_time = datetime.datetime.now()
    print("Time Taken : " + str(end_time - start_time))



start_time = datetime.datetime.now()

original_images_path = "Dataset/224/"
original_excel_path = 'Labels/Eyepacs(88457)_Splitted.xlsx'

sampling_images_path = "Dataset/ResizeCrop224/"
sampling_excel_path = 'Labels/Eyepacs(88457)_Splitted_ResizeCrop.xlsx'

# original_images_num = 88700
target_images_per_class = 17740
image_dimension = 224
train_percentage = 0.7
classes_num = 5
batch_size = 32
epochs = 5
train_2_times = True
classes_distribution_in_main_arrays = "01234"
sampling = 'resize_crop'
weights = 'imagenet'
optimizer = 'adam'
loss_function = 'sparse_categorical_crossentropy'
metrics = 'accuracy'

print("Original Images Number = 88700")
print("Image Dimension =", image_dimension)
print("Classes =", classes_num)
print("Train Percentage =", train_percentage)
print("Classes Distribution =", classes_distribution_in_main_arrays)
print("Sampling =", sampling)
print("Base Model = VGG16")
print("Weights =", weights)
print("Batch Size =", batch_size)
print("Epochs =", epochs)
print("Optimizer =", optimizer)
print("Loss Function =", loss_function)
print("Metrics =", metrics)
print("Train 2 Times =", train_2_times)

train_size_per_class_original = np.zeros(classes_num, int)
test_size_per_class_original = np.zeros(classes_num, int)

train_size_per_class_sampling = np.zeros(classes_num, int)
test_size_per_class_sampling = np.zeros(classes_num, int)

images_num_per_class_original = np.zeros(classes_num, int)
images_num_per_class_sampling = np.zeros(classes_num, int)

classes_counter = np.zeros(classes_num, int)

ws_original = getSheets(original_excel_path)
ws_sampling = getSheets(sampling_excel_path)

getNumOfImagesPerClass()
classes_images = readImagesToArrays()
classes_train_images_original, classes_test_images_original, classes_train_images_sampling, classes_test_images_sampling = splitData()

train_images = np.zeros(((sum(train_size_per_class_original) + sum(train_size_per_class_sampling)), image_dimension, image_dimension, 3), dtype='uint8')
train_labels = np.zeros(((sum(train_size_per_class_original) + sum(train_size_per_class_sampling)), 1), dtype='uint8')

test_images = np.zeros(((sum(test_size_per_class_original) + sum(test_size_per_class_sampling)), image_dimension, image_dimension, 3), dtype='uint8')
test_labels = np.zeros(((sum(test_size_per_class_original) + sum(test_size_per_class_sampling)), 1), dtype='uint8')

writeToMainArrays(classes_distribution_in_main_arrays)
checkArrays()
# model = createModel()
# model = trainModel(_model=model, _epochs=epochs, _train_2_times=train_2_times)
# evaluateModel(_model=model)

model_yesNo_path = "C:/Users/Youssef/PycharmProjects/Graduation/YesNo_2Classes_fullModel.h5"
model_stages_path = "C:/Users/Youssef/PycharmProjects/Graduation/Stages_4Classes_fullModel.h5"

model_yesNo = keras.models.load_model(model_yesNo_path)
model_stages = keras.models.load_model(model_stages_path)

total_test_size = sum(test_size_per_class_original) + sum(test_size_per_class_sampling)

test_predictions = np.zeros((total_test_size, 1), dtype='uint8')

prediction_yesNo = model_yesNo.predict(test_images)
prediction_yesNo_max = np.argmax(prediction_yesNo, axis=1)

prediction_stages = model_stages.predict(test_images)
prediction_stages_max = np.argmax(prediction_stages, axis=1)

for i in range(total_test_size):
    print("i = ", i)
    if prediction_yesNo_max[i] == 1:
        test_predictions[i] = prediction_stages_max[i] + 1
        print("Stages : ", test_predictions[i])
    else:
        test_predictions[i] = prediction_yesNo_max[i]
        print("YesNo : ", test_predictions[i])

print(confusion_matrix(test_labels, test_predictions))
# print(tf.math.confusion_matrix(labels=test_labels, predictions=test_predictions))
calculateTime()