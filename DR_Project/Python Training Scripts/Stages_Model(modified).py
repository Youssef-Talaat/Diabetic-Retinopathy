import numpy as np
from keras.layers import Dense, AveragePooling2D, GlobalAveragePooling2D, Dropout, Flatten
from keras.optimizers import SGD
from keras.applications import VGG16, VGG19, DenseNet201, ResNet50
from keras.models import Model
import cv2
import tensorflow as tf
from openpyxl import load_workbook
import datetime


def getNumOfImagesPerClass():
    print("getNumOfImagesPerClass()")
    for i in range(1, images_num + 1, 1):
        images_num_per_class[ws.cell(i, 2).value] += 1


def readImagesToArrays():
    print("readImagesToArrays()")
    if classes_num == 4:
        temp0 = np.zeros((images_num_per_class[1], image_dimension, image_dimension, 3), dtype='uint8')
        temp1 = np.zeros((images_num_per_class[2], image_dimension, image_dimension, 3), dtype='uint8')
        temp2 = np.zeros((images_num_per_class[3], image_dimension, image_dimension, 3), dtype='uint8')
        temp3 = np.zeros((images_num_per_class[4], image_dimension, image_dimension, 3), dtype='uint8')
        classes_images = np.array([temp0, temp1, temp2, temp3])
    elif classes_num == 5:
        temp0 = np.zeros((images_num_per_class[0], image_dimension, image_dimension, 3), dtype='uint8')
        temp1 = np.zeros((images_num_per_class[1], image_dimension, image_dimension, 3), dtype='uint8')
        temp2 = np.zeros((images_num_per_class[2], image_dimension, image_dimension, 3), dtype='uint8')
        temp3 = np.zeros((images_num_per_class[3], image_dimension, image_dimension, 3), dtype='uint8')
        temp4 = np.zeros((images_num_per_class[4], image_dimension, image_dimension, 3), dtype='uint8')
        classes_images = np.array([temp0, temp1, temp2, temp3, temp4])

    for i in range(1, images_num + 1, 1):
        '''if ws.cell(i, 2).value == 0 and classification_num == 2:
            classes_images[0][classes_counter[0]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
            classes_counter[0] += 1'''
        if ws.cell(i, 2).value == 1:
            classes_images[0][classes_counter[0]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
            classes_counter[0] += 1
        elif ws.cell(i, 2).value == 2:
            classes_images[1][classes_counter[1]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
            classes_counter[1] += 1
        elif ws.cell(i, 2).value == 3:
            classes_images[2][classes_counter[2]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
            classes_counter[2] += 1
        elif ws.cell(i, 2).value == 4:
            classes_images[3][classes_counter[3]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
            classes_counter[3] += 1

    print(images_num_per_class)
    return classes_images


def splitData():
    print("splitData()")
    for c in range(classes_num):
        train_size_per_class[c] = int(images_num_per_class[c + 1] * train_percentage + 0.5)
        test_size_per_class[c] = images_num_per_class[c + 1] - train_size_per_class[c]
        print(str(train_size_per_class[c]) + " , " + str(test_size_per_class[c]))

    classes_train_images_local = np.array([classes_images[0][0:train_size_per_class[0]],
                                           classes_images[1][0:train_size_per_class[1]],
                                           classes_images[2][0:train_size_per_class[2]],
                                           classes_images[3][0:train_size_per_class[3]]])
    classes_test_images_local = np.array([classes_images[0][train_size_per_class[0]:],
                                          classes_images[1][train_size_per_class[1]:],
                                          classes_images[2][train_size_per_class[2]:],
                                          classes_images[3][train_size_per_class[3]:]])
    '''for i in range(classes_num):
        classes_train_images[i] = classes_images[i][0:train_size_per_class[i]]
        classes_test_images[i] = classes_images[i][train_size_per_class[i]:]'''
    return classes_train_images_local, classes_test_images_local


def augmentPerClass(class_index):
    print("augmentPerClass(",class_index,")")
    angles = [90, 180, 270]
    train_target = np.amin(train_size_per_class) * (augmentation_per_image + 1)
    test_target = np.amin(test_size_per_class) * (augmentation_per_image + 1)
    train_diff = train_target - train_size_per_class[class_index]
    test_diff = test_target - test_size_per_class[class_index]
    if train_diff <= 0:
        index = class_index
        for i in range(train_target):
            train_images[index] = classes_train_images[class_index][i]
            train_labels[index] = class_index
            index += classes_num
        index = class_index
        for i in range(test_target):
            test_images[index] = classes_test_images[class_index][i]
            test_labels[index] = class_index
            index += classes_num
    else:
        whole_loop = train_diff // train_size_per_class[class_index]
        mod_train = train_diff % train_size_per_class[class_index]
        mod_test = test_diff % test_size_per_class[class_index]
        train_augment = np.zeros((train_diff, image_dimension, image_dimension, 3), dtype='uint8')
        test_augment = np.zeros((test_diff, image_dimension, image_dimension, 3), dtype='uint8')
        angle_index = 0
        train_counter = 0
        test_counter = 0
        print("Whole Loop :", whole_loop)
        print("Train Diff:", train_diff)
        print("Test Diff:", test_diff)
        print("Modules Train:", mod_train)
        print("Modules Test :", mod_test)
        for i in range(whole_loop):
            for j in range(train_size_per_class[class_index]):
                M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[i], 1.0)
                train_augment[train_counter] = cv2.warpAffine(classes_train_images[class_index][j], M, (image_dimension, image_dimension))
                train_counter += 1
            for j in range(test_size_per_class[class_index]):
                M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[i], 1.0)
                test_augment[test_counter] = cv2.warpAffine(classes_test_images[class_index][j], M, (image_dimension, image_dimension))
                test_counter += 1
            angle_index += 1
        for i in range(mod_train):
            M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[angle_index], 1.0)
            train_augment[train_counter] = cv2.warpAffine(classes_train_images[class_index][i], M, (image_dimension, image_dimension))
            train_counter += 1
        for i in range(mod_test):
            M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[angle_index], 1.0)
            test_augment[test_counter] = cv2.warpAffine(classes_test_images[class_index][i], M, (image_dimension, image_dimension))
            test_counter += 1
        #####################################################################################
        train_index = class_index
        test_index = class_index
        for i in range(train_size_per_class[class_index]):
            train_images[train_index] = classes_train_images[class_index][i]
            train_labels[train_index] = class_index
            train_index += classes_num
        for i in range(train_diff):
            train_images[train_index] = train_augment[i]
            train_labels[train_index] = class_index
            train_index += classes_num
        for i in range(test_size_per_class[class_index]):
            test_images[test_index] = classes_test_images[class_index][i]
            test_labels[test_index] = class_index
            test_index += classes_num
        for i in range(test_diff):
            test_images[test_index] = test_augment[i]
            test_labels[test_index] = class_index
            test_index += classes_num


def checkArrays():
    print("checkArrays()")
    zeros_array = np.zeros((image_dimension, image_dimension, 3), dtype='uint8')
    total_train = np.amin(train_size_per_class)*(augmentation_per_image + 1)*classes_num
    total_test = np.amin(test_size_per_class)*(augmentation_per_image + 1)*classes_num
    zero_train_counter = 0
    zero_test_counter = 0
    for i in range(total_train):
        if np.array_equal(train_images[i], zeros_array):
            zero_train_counter += 1
    for i in range(total_test):
        if np.array_equal(test_images[i], zeros_array):
            zero_test_counter += 1

    print("Zero Images in Train = ", zero_train_counter)
    print("Zero Images in Test = ", zero_test_counter)


def createModel():
    print("createModel()")
    baseModel = VGG16(include_top=False, weights='imagenet', input_tensor=None, input_shape=(image_dimension, image_dimension, 3))

    x = baseModel.output
    x = GlobalAveragePooling2D()(x)
    x = Dense(512, activation='relu')(x)
    x = Dropout(0.5)(x)
    headModel = Dense(classes_num, activation='softmax')(x)

    model = Model(inputs=baseModel.input, outputs=headModel)

    for layer in baseModel.layers:
        layer.trainable = False

    return model


def trainModel(model, epochs, top_layers_only):
    print("trainModel(model, epochs, top_layers_only)")
    model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])
    model.summary()
    model.fit(train_images, train_labels, epochs=epochs)

    if not top_layers_only:
        for layer in model.layers[:15]:
            layer.trainable = False
        for layer in model.layers[15:]:
            layer.trainable = True

        model.compile(optimizer=SGD(lr=0.0001, momentum=0.9), loss='sparse_categorical_crossentropy', metrics=['accuracy'])
        model.summary()
        model.fit(train_images, train_labels, epochs=epochs)

    return model


def evaluateModel(model):
    print("trainModel(model, epochs, top_layers_only)")
    # model.save('Stages_fullModel.h5')
    # model.save_weights('Stages_weights.h5')

    test_loss, test_acc = model.evaluate(test_images, test_labels)

    print('\nTest accuracy:', test_acc)
    print('\nTest loss:', test_loss)

    pred = model.predict(test_images)
    print(pred)
    max_predictions = tf.argmax(pred, axis=1)
    print(max_predictions)
    print(tf.math.confusion_matrix(labels=test_labels, predictions=max_predictions))

    print('Sum Test: ', sum(sum(pred)))


def calculateTime():
    print("trainModel(model, epochs, top_layers_only)")
    end_time = datetime.datetime.now()
    print("Time Taken : " + str(end_time - start_time))



start_time = datetime.datetime.now()

images_path = "C:/Users/Youssef/PycharmProjects/AlexNet/dataset_resized_2/"
excel_path = 'Labels/Eyepacs(88457)_Filtered.xlsx'
sheet_name = 'Sheet1'

images_num = 88702 - 245
image_dimension = 227
train_percentage = 0.7
classes_num = 4
augmentation_per_image = 3

test_percentage = 1 - train_percentage
wb = load_workbook(excel_path)
ws = wb[sheet_name]

train_size_per_class = np.zeros(classes_num, int)
test_size_per_class = np.zeros(classes_num, int)

classes_images = None
classes_train_images = None
classes_test_images = None

images_num_per_class = np.zeros(classes_num + 1, int)
classes_counter = np.zeros(classes_num, int)


getNumOfImagesPerClass()
classes_images = readImagesToArrays()
classes_train_images, classes_test_images = splitData()

train_images = np.zeros((np.amin(train_size_per_class)*(augmentation_per_image + 1)*classes_num, image_dimension, image_dimension, 3), dtype='uint8')
train_labels = np.zeros((np.amin(train_size_per_class)*(augmentation_per_image + 1)*classes_num, 1), dtype='uint8')

test_images = np.zeros((np.amin(test_size_per_class)*(augmentation_per_image + 1)*classes_num, image_dimension, image_dimension, 3), dtype='uint8')
test_labels = np.zeros((np.amin(test_size_per_class)*(augmentation_per_image + 1)*classes_num, 1), dtype='uint8')

for c in range(classes_num):
    augmentPerClass(c)
checkArrays()
model = createModel()
model = trainModel(model=model, epochs=5, top_layers_only=False)
evaluateModel(model=model)
calculateTime()