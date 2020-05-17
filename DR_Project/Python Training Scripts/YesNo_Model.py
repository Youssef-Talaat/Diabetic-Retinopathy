import numpy as np
from keras.layers import Dense, AveragePooling2D, GlobalAveragePooling2D, Dropout, Flatten
from keras.applications import VGG16, VGG19, DenseNet201, ResNet50
from keras.models import Model
from keras.optimizers import SGD
import cv2
import tensorflow as tf
from openpyxl import load_workbook
import datetime


def add_to_main_arrays(main_i, main_l, temp, temp_label, temp_size, index):
    i = 0
    for i in range(temp_size):
        main_i[index] = temp[i]
        main_l[index] = temp_label
        index += 1
    return index


start_time = datetime.datetime.now()

images_no = 88702 - 245
image_dimension = 224
train_percentage = 0.7
classes_num = 5
images_path = "dataset_resized_2/"

classes = np.zeros(classes_num, int)
classes_counter = np.zeros(classes_num, int)

class_train_size = np.zeros(classes_num, int)
class_test_size = np.zeros(classes_num, int)

wb = load_workbook('Labels_2.xlsx')
ws = wb['Sheet 1']
cols = 1

for i in range(1, images_no + 1, 1):
    classes[ws.cell(i, 2).value] += 1

class_I0 = np.zeros((classes[0], image_dimension, image_dimension, 3), dtype='uint8')
class_I1 = np.zeros((classes[1], image_dimension, image_dimension, 3), dtype='uint8')
class_I2 = np.zeros((classes[2], image_dimension, image_dimension, 3), dtype='uint8')
class_I3 = np.zeros((classes[3], image_dimension, image_dimension, 3), dtype='uint8')
class_I4 = np.zeros((classes[4], image_dimension, image_dimension, 3), dtype='uint8')

for i in range(1, images_no + 1, 1):
    if ws.cell(i, 2).value == 0:
        class_I0[classes_counter[0]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
        classes_counter[0] += 1
    elif ws.cell(i, 2).value == 1:
        class_I1[classes_counter[1]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
        classes_counter[1] += 1
    elif ws.cell(i, 2).value == 2:
        class_I2[classes_counter[2]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
        classes_counter[2] += 1
    elif ws.cell(i, 2).value == 3:
        class_I3[classes_counter[3]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
        classes_counter[3] += 1
    elif ws.cell(i, 2).value == 4:
        class_I4[classes_counter[4]] = cv2.imread(images_path + str(ws.cell(i, 1).value))
        classes_counter[4] += 1

print(classes)

for c in range(classes.size):
    class_train_size[c] = int(classes[c] * train_percentage + 0.5)
    class_test_size[c] = classes[c] - class_train_size[c]
    print(str(class_train_size[c]) + " , " + str(class_test_size[c]))

train_number = sum(class_train_size)
test_number = sum(class_test_size)

train_images = np.zeros((42592, image_dimension, image_dimension, 3), dtype='uint8')
train_labels = np.zeros((42592, 1), dtype='uint8')

test_images = np.zeros((18240, image_dimension, image_dimension, 3), dtype='uint8')
test_labels = np.zeros((18240, 1), dtype='uint8')

print("Train Number : " + str(train_number))
print("Test Number : " + str(test_number))

c0_train = class_I0[0:class_train_size[0]]
c0_test = class_I0[class_train_size[0]:]
c1_train = class_I1[0:class_train_size[1]]
c1_test = class_I1[class_train_size[1]:]
c2_train = class_I2[0:class_train_size[2]]
c2_test = class_I2[class_train_size[2]:]
c3_train = class_I3[0:class_train_size[3]]
c3_test = class_I3[class_train_size[3]:]
c4_train = class_I4[0:class_train_size[4]]
c4_test = class_I4[class_train_size[4]:]

#####################################################################################
#####################################################################################
cYes_train_all = np.zeros((21296, image_dimension, image_dimension, 3), dtype='uint8')
cYes_test_all = np.zeros((9120, image_dimension, image_dimension, 3), dtype='uint8')

# Classes except 0 :

train_index = 0
test_index = 0
angles = [90, 180, 270]

c1_train_all = np.zeros((5324, image_dimension, image_dimension, 3), dtype='uint8')
c2_train_all = np.zeros((5324, image_dimension, image_dimension, 3), dtype='uint8')
c3_train_all = np.zeros((5324, image_dimension, image_dimension, 3), dtype='uint8')
c4_train_all = np.zeros((5324, image_dimension, image_dimension, 3), dtype='uint8')

c1_test_all = np.zeros((2280, image_dimension, image_dimension, 3), dtype='uint8')
c2_test_all = np.zeros((2280, image_dimension, image_dimension, 3), dtype='uint8')
c3_test_all = np.zeros((2280, image_dimension, image_dimension, 3), dtype='uint8')
c4_test_all = np.zeros((2280, image_dimension, image_dimension, 3), dtype='uint8')

#####################################################################################
#####################################################################################
# Class 1 :
c1_train_new = np.zeros((5324 - 4333, image_dimension, image_dimension, 3), dtype='uint8')
c1_test_new = np.zeros((2280 - 1857, image_dimension, image_dimension, 3), dtype='uint8')

counter = 0
for i in range(4333,5324,1):
    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[i%3], 1.0)
    c1_train_new[counter] = cv2.warpAffine(c1_train[counter], M, (image_dimension, image_dimension))
    counter += 1

counter = 0
for i in range(1857,2280,1):
    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[i%3], 1.0)
    c1_test_new[counter] = cv2.warpAffine(c1_test[counter], M, (image_dimension, image_dimension))
    counter += 1

counter = 0
for i in range(4333):
    c1_train_all[counter] = c1_train[i]
    counter += 1
for i in range(991):
    c1_train_all[counter] = c1_train_new[i]
    counter += 1

counter = 0
for i in range(1857):
    c1_test_all[counter] = c1_test[i]
    counter += 1
for i in range(423):
    c1_test_all[counter] = c1_test_new[i]
    counter += 1

#####################################################################################
#####################################################################################
# Class 2 :

c2_train_all = c2_train[0:5324]
c2_test_all = c2_test[0:2280]

#####################################################################################
#####################################################################################
# Class 3 :
c3_train_new = np.zeros((5324 - 1458, image_dimension, image_dimension, 3), dtype='uint8')
c3_test_new = np.zeros((2280 - 625, image_dimension, image_dimension, 3), dtype='uint8')

counter = 0
for i in range(1458, 5354 - 3, 3):
    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[0], 1.0)
    c3_train_new[counter] = cv2.warpAffine(c3_train[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[1], 1.0)
    c3_train_new[counter + 1] = cv2.warpAffine(c3_train[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[2], 1.0)
    c3_train_new[counter + 2] = cv2.warpAffine(c3_train[counter], M, (image_dimension, image_dimension))

    counter += 1

counter = 0
for i in range(625, 2280 - 3, 3):
    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[0], 1.0)
    c3_test_new[counter] = cv2.warpAffine(c3_test[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[1], 1.0)
    c3_test_new[counter + 1] = cv2.warpAffine(c3_test[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[2], 1.0)
    c3_test_new[counter + 2] = cv2.warpAffine(c3_test[counter], M, (image_dimension, image_dimension))

    counter += 1

counter = 0
for i in range(1458):
    c3_train_all[counter] = c3_train[i]
    counter += 1
for i in range(3866):
    c3_train_all[counter] = c3_train_new[i]
    counter += 1

counter = 0
for i in range(625):
    c3_test_all[counter] = c3_test[i]
    counter += 1
for i in range(1655):
    c3_test_all[counter] = c3_test_new[i]
    counter += 1


#####################################################################################
#####################################################################################
# Class 4 :
c4_train_new = np.zeros((5324 - 1331, image_dimension, image_dimension, 3), dtype='uint8')
c4_test_new = np.zeros((2280 - 570, image_dimension, image_dimension, 3), dtype='uint8')

counter = 0
for i in range(1331, 5324, 3):
    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[0], 1.0)
    c4_train_new[counter] = cv2.warpAffine(c4_train[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[1], 1.0)
    c4_train_new[counter + 1] = cv2.warpAffine(c4_train[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[2], 1.0)
    c4_train_new[counter + 2] = cv2.warpAffine(c4_train[counter], M, (image_dimension, image_dimension))

    counter += 1

counter = 0
for i in range(570, 2280, 3):
    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[0], 1.0)
    c4_test_new[counter] = cv2.warpAffine(c4_test[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[1], 1.0)
    c4_test_new[counter + 1] = cv2.warpAffine(c4_test[counter], M, (image_dimension, image_dimension))

    M = cv2.getRotationMatrix2D((image_dimension / 2, image_dimension / 2), angles[2], 1.0)
    c4_test_new[counter + 2] = cv2.warpAffine(c4_test[counter], M, (image_dimension, image_dimension))

    counter += 1

counter = 0
for i in range(1331):
    c4_train_all[counter] = c4_train[i]
    counter += 1
for i in range(3993):
    c4_train_all[counter] = c4_train_new[i]
    counter += 1

counter = 0
for i in range(570):
    c4_test_all[counter] = c4_test[i]
    counter += 1
for i in range(1710):
    c4_test_all[counter] = c4_test_new[i]
    counter += 1

#####################################################################################
#####################################################################################
cYes_train_all[0:5324] = c1_train_all
cYes_train_all[5324:10648] = c2_train_all
cYes_train_all[10648:15972] = c3_train_all
cYes_train_all[15972:21296] = c4_train_all
cYes_test_all[0:2280] = c1_test_all
cYes_test_all[2280:4560] = c2_test_all
cYes_test_all[4560:6840] = c3_test_all
cYes_test_all[6840:9120] = c4_test_all

index = 0
for i in range(21296):
    train_images[index] = c0_train[i]
    train_labels[index] = 0
    index += 1
    train_images[index] = cYes_train_all[i]
    train_labels[index] = 1
    index += 1

index = 0
for i in range(9120):
    test_images[index] = c0_test[i]
    test_labels[index] = 0
    index += 1
    test_images[index] = cYes_test_all[i]
    test_labels[index] = 1
    index += 1

print("Train Labels", train_labels)
print("Test Labels", test_labels)

print(train_images.size + test_images.size)

print(c0_train.size + c0_test.size +
      cYes_train_all.size + cYes_test_all.size)

#####################################################################################
#####################################################################################
baseModel = VGG16(include_top=False, weights='imagenet', input_tensor=None, input_shape=(image_dimension, image_dimension, 3))

'''headModel = baseModel.output
headModel = AveragePooling2D(pool_size=(4, 4))(headModel)
headModel = Flatten(name='flatten')(headModel)
headModel = Dense(128, activation='relu')(headModel)
headModel = Dropout(0.3)(headModel)
headModel = Dense(2, activation='softmax')(headModel)'''

x = baseModel.output
x = GlobalAveragePooling2D()(x)
x = Dense(512, activation='relu')(x)
x = Dropout(0.5)(x)
headModel = Dense(2, activation='softmax')(x)

model = Model(inputs=baseModel.input, outputs=headModel)

for layer in baseModel.layers:
   layer.trainable = False

model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])
model.summary()
model.fit(train_images, train_labels, epochs=5)

for layer in model.layers[:15]:
   layer.trainable = False
for layer in model.layers[15:]:
   layer.trainable = True

model.compile(optimizer=SGD(lr=0.0001, momentum=0.9), loss='sparse_categorical_crossentropy', metrics=['accuracy'])
model.summary()
model.fit(train_images, train_labels, epochs=5)

#####################################################################################
#####################################################################################
model.save('YesNo_fullModel.h5')
model.save_weights('YesNo_weights.h5')

test_loss, test_acc = model.evaluate(test_images, test_labels)

print('\nTest accuracy:', test_acc)
print('\nTest loss:', test_loss)

pred = model.predict(test_images)
print(pred)
max_predictions = tf.argmax(pred, axis=1)
print(max_predictions)
print(tf.math.confusion_matrix(labels=test_labels, predictions=max_predictions))

print('Sum Test: ', sum(sum(pred)))

end_time = datetime.datetime.now()
print("Time Taken : " + str(end_time - start_time))
